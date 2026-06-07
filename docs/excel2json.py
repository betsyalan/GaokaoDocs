#!/usr/bin/env python3
"""
高考志愿表 Excel → JSON 转换脚本
用法: python3 docs/excel2json.py
效果: 读取 docs/*志愿表*.xlsx → 输出 docs/volunteer-data.json
"""
import openpyxl, json, re, os

DOCS_DIR = os.path.dirname(os.path.abspath(__file__))

def parse_year_data(row):
    """解析年份数据（col22-45，4年 × 6字段）"""
    years = ['2025物理', '2024物理', '2023物理', '2022物理']
    fields = ['录取人数', '线差', '最低分', '最低位次', '等效位差', '等效分差']
    result = {}
    for yi, year in enumerate(years):
        base = 21 + yi * 6  # 0-indexed col22-27, 28-33, 34-39, 40-45
        obj = {}
        for fi, fname in enumerate(fields):
            val = row[base + fi]
            if val not in [None, '', '-']:
                obj[fname] = str(val)
        if obj:
            result[year] = obj
    return result

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    groups = []
    current_group = None
    last_group_num = None

    for row in rows:
        group_num = row[0]
        is_new_group = group_num is not None and group_num != '' and group_num != last_group_num

        if is_new_group:
            if current_group:
                groups.append(current_group)

            raw_name = str(row[3] or '')
            school_name = raw_name.split('---')[0] if '---' in raw_name else (raw_name or str(row[17] or '').strip())
            level_tags = [t.strip() for t in (str(row[4] or '')).split('\n') if t.strip()]

            current_group = {
                'group_num': len(groups) + 1,
                'school_code': str(row[2] or ''),
                'school_name': school_name,
                'group_probability': str(row[1] or ''),
                'group_probability_raw': str(row[1] or ''),
                'level_tags': level_tags,
                'location': str(row[5] or ''),
                'type_label': str(row[6] or ''),
                'nature': str(row[7] or ''),
                'ranking': str(row[8] or ''),
                'master_count': str(row[9] or ''),
                'doctor_count': str(row[10] or ''),
                'grad_rate': str(row[11] or ''),
                'postgrad_rate': str(row[12] or ''),
                'employ_rate': str(row[13] or ''),
                'subject_req': str(row[14] or ''),
                'group_plan': str(row[18] or ''),
                'majors': [],
                'group_years': parse_year_data(row)
            }
            last_group_num = group_num
        elif current_group:
            major_name = str(row[17] or '').strip()
            if major_name:
                major = {
                    'code': str(row[16] or ''),
                    'name': major_name.replace('\n', ' ').replace('\r', ''),  # 清理单元格内换行符
                    'probability': str(row[15] or ''),
                    'plan': str(row[18] or ''),
                    'tuition': row[19] if row[19] not in [None, '', '-'] else None,
                    'duration': str(row[20] or ''),
                    'years': parse_year_data(row)
                }
                m = re.search(r'\s([A-D][+-]?)\s', major_name)
                if m:
                    major['rating'] = m.group(1)
                current_group['majors'].append(major)

    if current_group:
        groups.append(current_group)

    score_match = re.search(r'(\d+分)', filepath)
    score = score_match.group(1) if score_match else 'unknown'

    meta_raw = str(list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0][0] or '')
    parts = meta_raw.split()

    return {
        score: {
            'meta': {
                'province': parts[1] if len(parts) > 1 else '',
                'gender': parts[3] if len(parts) > 3 else '',
                'level': parts[4] if len(parts) > 4 else '',
                'score': parts[5] if len(parts) > 5 else '',
                'subjects': parts[7] if len(parts) > 7 else ''
            },
            'groups': groups
        }
    }

if __name__ == '__main__':
    result = {}
    for f in sorted(os.listdir(DOCS_DIR)):
        if '志愿表' in f and f.endswith('.xlsx'):
            print(f'解析: {f}')
            data = parse_excel(os.path.join(DOCS_DIR, f))
            result.update(data)

    outpath = os.path.join(DOCS_DIR, 'volunteer-data.json')
    with open(outpath, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    for k, v in result.items():
        total = len(v['groups'])
        majors = sum(len(g['majors']) for g in v['groups'])
        print(f'  {k}: {total}组 / {majors}专业')

    print(f'\n✅ 已导出 {outpath}')
